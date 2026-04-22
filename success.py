import streamlit as st
from fpdf import FPDF
import pandas as pd
import os
import openpyxl
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from dotenv import load_dotenv

def show_success():
    st.title("🎉 Order Confirmed!")
    st.session_state.invoice_generated = True
    if "order_saved" not in st.session_state:
        st.session_state.order_saved = False
    if "email_sent" not in st.session_state:
        st.session_state.email_sent = False

    if st.session_state.get("invoice_generated", False):
        # Convert cart to DataFrame
        invoice_df = pd.DataFrame(st.session_state.cart)
        invoice_df['Subtotal'] = invoice_df['Price'] * invoice_df['qty']
        invoice_df = invoice_df[['Name', 'qty', 'Price', 'Subtotal']]

        st.markdown("### 🧾 Invoice")
        st.dataframe(invoice_df, use_container_width=True,hide_index=True)
        st.write(f"Total Paid: ₹{st.session_state.order['total']}")
        # Convert DataFrame to CSV
        csv = invoice_df.to_csv(index=False).encode('utf-8')
        order_id = f"ORD-{pd.Timestamp.now().strftime('%Y%m%d%H%M%S')}"
        invoice_df['order_id'] = order_id
        val_timestamp = invoice_df['timestamp'] = datetime.now().strftime('%Y-%m-%d-%H:%M:%S')
        st.markdown(f" Order-Id {order_id} - {val_timestamp}")
        customer_id = st.session_state.order['phone'] +'_'+ val_timestamp
        st.session_state.order.update({"customer_id": customer_id})
        st.session_state.order.update({"order_id": order_id})
        st.session_state.order.update({"order_date": val_timestamp})
        customer_df = pd.DataFrame([st.session_state.order])
        orders_df = pd.DataFrame([{
            "order_id": st.session_state.order.get("order_id"),
            "customer_id": st.session_state.order.get("customer_id"),
            "total": st.session_state.order.get("total"),
            "order_date": st.session_state.order.get("order_date"),
            "payment":st.session_state.order.get("payment"),
            "payment_status":st.session_state.order.get("payment_status")
        }])
        #st.write(st.session_state.order)
        # Pdf creation
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        # Title
        pdf.cell(0, 10, "Invoice", ln=True, align='C')
        pdf.ln(5)

        # ---- TABLE HEADER ----
        col_widths = [60, 20, 40, 40]  # adjust widths

        headers = ["Product", "Qty", "Price", "Subtotal"]

        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, border=1, align='C')
        pdf.ln()

        # ---- TABLE DATA ----
        total = 0

        for item in st.session_state.cart:
            subtotal = item["Price"] * item["qty"]
            total += subtotal

            pdf.cell(col_widths[0], 10, item["Name"], border=1)
            pdf.cell(col_widths[1], 10, str(item["qty"]), border=1, align='C')
            pdf.cell(col_widths[2], 10, f"Rs.{item['Price']}", border=1, align='R')
            pdf.cell(col_widths[3], 10, f"Rs.{subtotal}", border=1, align='R')
            pdf.ln()

        # ---- TOTAL ROW ----
        pdf.set_font('Arial', '', 12)
        pdf.cell(sum(col_widths[:-1]), 10, "Total", border=1, align='R')
        pdf.cell(col_widths[-1], 10, f"Rs.{total}", border=1, align='R')

        # ---- EXPORT ----
        pdf_bytes = pdf.output(dest='S').encode('latin1')

        if not st.session_state.order_saved:
            save_order_details(invoice_df)
            customer_details(customer_df)
            order_details(orders_df)
            st.session_state.order_saved = True
            # ✅ Send email
            email_sent = send_email(
                st.session_state.order.get("email"),
                st.session_state.order,invoice_df, pdf_bytes
            )

            if email_sent:
                st.success("📧 Confirmation email sent!")
                st.session_state.email_sent = True
            else:
                st.warning("Email failed to send")

        c1, c2, c3 = st.columns([1,1,1])

        with c1:
            st.download_button(
                label="📥 Download Invoice (CSV)",
                data=csv,
                file_name="invoice.csv",
                mime="text/csv"
            )

        with c2:
            st.download_button(
                label="📥 Download Invoice (PDF)",
                data=pdf_bytes,
                file_name="invoice.pdf",
                mime="application/pdf"
            )
        with c3:
            if st.button("🛍 Back to Shop"):
                st.session_state.cart.clear()
                st.session_state.pop("order", None)
                st.session_state.order_saved = False
                st.session_state.page = "shop"
                st.session_state.clear()
                st.rerun()

def save_order_details(new_order, file="Data/orders_details.xlsx"):
    try:
        if "order_id" in new_order.columns:
            existing_df = pd.read_excel(file)
            final_df = pd.concat([existing_df, new_order], ignore_index=True)
            if new_order["order_id"].iloc[0] in existing_df["order_id"].values:
                return  # already saved
    except FileNotFoundError:
        final_df = new_order
    final_df.to_excel(file, index=False)
    st.success("Order Placed ✅")

def customer_details(new_customer, file="Data/Customer.xlsx"):
    new_customer = new_customer.drop(columns=["items"])
    try:
        if "customer_id" in new_customer.columns:
            existing_df = pd.read_excel(file)
            final_df = pd.concat([existing_df, new_customer], ignore_index=True)
            if new_customer["customer_id"].iloc[0] in existing_df["customer_id"].values:
                return  # already saved
    except FileNotFoundError:
        final_df = new_customer
    final_df.to_excel(file, index=False)
    st.success("Customer Details Saved ✅")

def order_details(order_que, file="Data/orders.xlsx"):
    try:
        if "order_id" in order_que.columns:
            existing_df = pd.read_excel(file)
            final_df = pd.concat([existing_df, order_que], ignore_index=True)
            if order_que["order_id"].iloc[0] in existing_df["order_id"].values:
                return  # already saved
    except FileNotFoundError:
        final_df = order_que
    final_df.to_excel(file, index=False)
    st.success("Order Que Added ✅")



def send_email(to_email, order_details,invoice_df, pdf_bytes):
    load_dotenv()

    sender_email = st.secrets["EMAIL_USER"]
    app_password = st.secrets["EMAIL_PASS"]

    subject = "Order Confirmation ✅"
    #df = pd.DataFrame([invoice_df])
    html_table = invoice_df.to_html(index=False)
    body = f"""
    <h2>Order Confirmed 🎉</h2>
    <p><b>Order ID:</b> {order_details.get('order_id')}</p>
    <p><b>Name:</b> {order_details.get('name')}</p>
    <p><b>Total:</b> ₹{order_details.get('total')}</p>
    <p>
    {html_table}
    </p>
    <h2>Thank you for your Order 🎉</h2>
    
    """

    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "html"))


    #pdf_bytes = pdf.output(dest='S').encode('latin1')
    part = MIMEApplication(pdf_bytes, _subtype="pdf")
    part.add_header('Content-Disposition', 'attachment', filename="invoice.pdf")
    msg.attach(part)

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587,timeout=30)
        server.starttls()
        server.login(sender_email, app_password)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(e)
        return False